import pandas as pd
from openpyxl.styles import PatternFill


def save_dataframe_to_csv(dataframe, output_path):
    dataframe.to_csv(output_path, index=False)


def save_dataframe_to_xlsx(dataframe, output_path):
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    dataframe.to_excel(writer, sheet_name='Sheet1', index=False)
    worksheet = writer.sheets['Sheet1']
    create_alternating_colors(worksheet=worksheet, row_count=len(dataframe), color1="FFFFFF", color2="B8CCE4")
    set_header_color(worksheet, 'FFA500')
    workbook = writer.book
    workbook.save(output_path)


def create_alternating_colors(worksheet, row_count, color1, color2):
    for row in worksheet.iter_rows(min_row=2, max_row=row_count + 1):
        for cell in row:
            cell.fill = PatternFill(start_color=color1 if cell.row % 2 == 0 else color2,
                                    end_color=color1 if cell.row % 2 == 0 else color2, fill_type="solid")


def set_header_color(worksheet, color):
    for cell in worksheet[1]:
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
